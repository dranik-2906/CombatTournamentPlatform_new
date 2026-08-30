import logging
import os
import tempfile
import uuid
import io
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.db.models import Q, Count
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from users.decorators import tournament_admin_required
from ..models import (
    Tournament, Fighter, TournamentRegistration,
    AgeWeightCategory, AgeGroup, WeightCategory,
    TournamentCheckpoint, RegistrationCheckpoint
)
from ..services import CategoryService
from ..forms import FighterForm, ExcelUploadForm, TournamentCheckpointForm

logger = logging.getLogger('tournaments')


@login_required
@tournament_admin_required
def participants_management(request):
    """Управление участниками"""
    tournaments = Tournament.objects.all().order_by('-start_date')
    selected_id = request.GET.get('tournament')
    selected = None
    registrations = TournamentRegistration.objects.none()

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    sort = request.GET.get('sort', 'fighter__last_name')

    age_weight_categories = AgeWeightCategory.objects.none()
    weight_categories = WeightCategory.objects.none()
    clubs = []

    if selected_id:
        selected = get_object_or_404(Tournament, pk=selected_id)
        registrations = selected.registrations.select_related(
            'fighter', 'age_weight_category'
        ).prefetch_related('fighter__user', 'checkpoint_statuses', 'checkpoint_statuses__checkpoint')

        age_weight_categories = selected.age_weight_categories.all()
        weight_categories = WeightCategory.objects.filter(tournament=selected)

        if search:
            registrations = registrations.filter(
                Q(fighter__first_name__icontains=search) |
                Q(fighter__last_name__icontains=search) |
                Q(fighter__club__icontains=search)
            )

        if status == 'approved':
            registrations = registrations.filter(is_approved=True)
        elif status == 'pending':
            registrations = registrations.filter(is_approved=False)

        registrations = registrations.order_by(sort)

        clubs = registrations.exclude(
            fighter__club__isnull=True
        ).exclude(fighter__club='').values_list(
            'fighter__club', flat=True
        ).distinct().order_by('fighter__club')

    stats = {}
    if selected:
        total = registrations.count()
        stats = {
            'total': total,
            'approved': registrations.filter(is_approved=True).count(),
            'pending': registrations.filter(is_approved=False).count(),
            'with_categories': registrations.filter(age_weight_category__isnull=False).count(),
        }

    return render(request, 'tournaments/participants_management.html', {
        'tournaments': tournaments,
        'selected': selected,
        'registrations': registrations,
        'stats': stats,
        'search': search,
        'status_filter': status,
        'sort': sort,
        'clubs': clubs,
        'weight_categories': weight_categories,
        'age_weight_categories': age_weight_categories,
    })


@login_required
@tournament_admin_required
def add_participant(request, tournament_id):
    tournament = get_object_or_404(Tournament, pk=tournament_id)
    if request.method == 'POST':
        form = FighterForm(request.POST, request.FILES)
        if form.is_valid():
            # Автоматически создаём пользователя для бойца
            from django.contrib.auth.models import User
            import random
            first = form.cleaned_data['first_name']
            last = form.cleaned_data['last_name']
            username = f"fighter_{first.lower()}_{last.lower()}_{random.randint(1000, 9999)}"
            user = User.objects.create_user(
                username=username,
                first_name=first,
                last_name=last,
                email=f"{username}@local.local",
                password='fighter123'
            )
            fighter = form.save(commit=False)
            fighter.user = user
            fighter.save()

            TournamentRegistration.objects.create(
                tournament=tournament,
                fighter=fighter,
            )
            messages.success(request, f'Боец {fighter.full_name} добавлен и зарегистрирован на турнир')
            return redirect('tournaments:participants_management')
    else:
        form = FighterForm()
    return render(request, 'tournaments/add_participant.html', {
        'form': form, 'tournament': tournament
    })

@login_required
@tournament_admin_required
def update_registration(request, registration_id):
    """Обновление статуса регистрации"""
    reg = get_object_or_404(
        TournamentRegistration.objects.select_related('fighter', 'tournament'),
        pk=registration_id
    )
    tournament_id = reg.tournament.id

    if request.method == 'POST':
        fighter = reg.fighter
        for field in ['first_name', 'last_name', 'club', 'coach']:
            value = request.POST.get(field, '').strip()
            if value:
                setattr(fighter, field, value)

        # === ИСПРАВЛЕНИЕ: явный парсинг даты из строки ===
        dob_str = request.POST.get('date_of_birth', '').strip()
        if dob_str:
            try:
                fighter.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
            except ValueError:
                try:
                    fighter.date_of_birth = datetime.strptime(dob_str, '%d.%m.%Y').date()
                except ValueError:
                    pass

        if request.POST.get('weight'):
            try:
                fighter.weight = float(request.POST.get('weight'))
            except ValueError:
                pass
        if request.POST.get('gender'):
            fighter.gender = request.POST.get('gender')
        fighter.save()

        # Обновляем динамические чекпоинты
        for cp in reg.tournament.checkpoints.all():
            field_name = f"checkpoint_{cp.id}"
            is_checked = field_name in request.POST
            RegistrationCheckpoint.objects.update_or_create(
                registration=reg,
                checkpoint=cp,
                defaults={'is_checked': is_checked}
            )

        reg = TournamentRegistration.objects.prefetch_related(
            'checkpoint_statuses', 'checkpoint_statuses__checkpoint'
        ).get(pk=reg.pk)

        was_approved = reg.is_approved
        reg.is_approved = reg.all_checks_passed

        if reg.is_approved and not was_approved:
            if fighter.date_of_birth and fighter.weight:
                category = CategoryService.find_category(
                    tournament=reg.tournament,
                    gender=fighter.gender,
                    birth_year=fighter.date_of_birth.year,
                    weight=float(fighter.weight)
                )
                if category:
                    reg.age_weight_category = category
                    messages.info(request, f'Автоматически назначена категория: {category.name}')
                else:
                    messages.warning(request, 'Не удалось подобрать категорию автоматически')

        awc_id = request.POST.get('age_weight_category')
        if awc_id:
            try:
                cat = AgeWeightCategory.objects.get(pk=awc_id, tournament=reg.tournament)
                reg.age_weight_category = cat
            except AgeWeightCategory.DoesNotExist:
                pass
        else:
            if not reg.age_weight_category:
                reg.age_weight_category = None

        reg.save()
        status_text = 'Допущен' if reg.is_approved else 'Не допущен'
        messages.success(request, f'Данные {fighter.full_name} обновлены. Статус: {status_text}')

    return redirect(f"{reverse('tournaments:participants_management')}?tournament={tournament_id}")


@login_required
@tournament_admin_required
def delete_registration(request, registration_id):
    """Удаление регистрации участника из турнира"""
    reg = get_object_or_404(
        TournamentRegistration.objects.select_related('fighter', 'tournament'),
        pk=registration_id
    )
    tournament_id = reg.tournament.id
    fighter_name = reg.fighter.full_name

    reg.delete()
    messages.success(request, f'Участник {fighter_name} удалён из турнира')
    return redirect(f"{reverse('tournaments:participants_management')}?tournament={tournament_id}")


@login_required
@tournament_admin_required
def auto_assign_categories_view(request, tournament_id):
    """Автоматическое распределение по категориям (только для допущенных)"""
    tournament = get_object_or_404(Tournament, pk=tournament_id)

    approved_without_cat = tournament.registrations.filter(
        is_approved=True, age_weight_category__isnull=True
    ).select_related('fighter')

    assigned = 0
    skipped = 0
    failed = 0
    errors = []

    for reg in approved_without_cat:
        fighter = reg.fighter
        if fighter.date_of_birth and fighter.weight:
            category = CategoryService.find_category(
                tournament, fighter.gender,
                fighter.date_of_birth.year, float(fighter.weight)
            )
            if category:
                reg.age_weight_category = category
                reg.save(update_fields=['age_weight_category'])
                assigned += 1
            else:
                failed += 1
                errors.append(f"{fighter.full_name}: не найдена подходящая категория")
        else:
            failed += 1
            errors.append(f"{fighter.full_name}: не указана дата рождения или вес")

    skipped = tournament.registrations.filter(
        age_weight_category__isnull=False
    ).count()

    if assigned > 0:
        messages.success(request, f"Распределено {assigned} участников")
    if skipped > 0:
        messages.info(request, f"Пропущено {skipped} (уже имеют категорию)")
    if failed > 0:
        messages.warning(request, f"Не удалось распределить {failed} участников")
    for error in errors[:5]:
        messages.warning(request, error)

    return redirect(f"{reverse('tournaments:participants_management')}?tournament={tournament_id}")


# ==================== ЧЕКПОИНТЫ ====================

@login_required
@tournament_admin_required
def manage_checkpoints(request, tournament_id):
    """Управление чекпоинтами турнира"""
    tournament = get_object_or_404(Tournament, pk=tournament_id)
    checkpoints = tournament.checkpoints.all()

    if request.method == 'POST':
        form = TournamentCheckpointForm(request.POST)
        if form.is_valid():
            cp = form.save(commit=False)
            cp.tournament = tournament
            cp.save()
            for reg in tournament.registrations.all():
                RegistrationCheckpoint.objects.get_or_create(
                    registration=reg, checkpoint=cp, defaults={'is_checked': False}
                )
            messages.success(request, f'Чекпоинт «{cp.name}» добавлен')
            return redirect('tournaments:manage_checkpoints', tournament_id=tournament_id)
    else:
        form = TournamentCheckpointForm()

    return render(request, 'tournaments/manage_checkpoints.html', {
        'tournament': tournament,
        'checkpoints': checkpoints,
        'form': form,
    })


@login_required
@tournament_admin_required
def edit_checkpoint(request, checkpoint_id):
    """Редактирование чекпоинта"""
    cp = get_object_or_404(TournamentCheckpoint, pk=checkpoint_id)
    tournament_id = cp.tournament.id

    if request.method == 'POST':
        form = TournamentCheckpointForm(request.POST, instance=cp)
        if form.is_valid():
            form.save()
            messages.success(request, f'Чекпоинт «{cp.name}» обновлён')
            return redirect('tournaments:manage_checkpoints', tournament_id=tournament_id)
    else:
        form = TournamentCheckpointForm(instance=cp)

    return render(request, 'tournaments/edit_checkpoint.html', {
        'form': form, 'checkpoint': cp, 'tournament': cp.tournament
    })


@login_required
@tournament_admin_required
def delete_checkpoint(request, checkpoint_id):
    """Удаление чекпоинта"""
    cp = get_object_or_404(TournamentCheckpoint, pk=checkpoint_id)
    tournament_id = cp.tournament.id
    name = cp.name
    cp.delete()
    messages.success(request, f'Чекпоинт «{name}» удалён')
    return redirect('tournaments:manage_checkpoints', tournament_id=tournament_id)


# ==================== EXCEL ====================

@login_required
@tournament_admin_required
def download_excel_template(request, tournament_id):
    """Скачать шаблон Excel для импорта участников"""
    tournament = get_object_or_404(Tournament, pk=tournament_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    headers = ['Имя', 'Фамилия', 'Дата рождения (ДД.ММ.ГГГГ)', 'Пол (М/Ж)', 'Вес (кг)', 'Клуб', 'Тренер']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    example = ['Иван', 'Иванов', '29.06.2005', 'М', '65.5', 'Боец Москва', 'Петров И.И.']
    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row=4, column=1, value="Правила заполнения:").font = Font(bold=True)
    ws.cell(row=5, column=1, value="• Пол: только 'М' (мужской) или 'Ж' (женский)")
    ws.cell(row=6, column=1, value="• Дата рождения: строго в формате ДД.ММ.ГГГГ, например 29.06.2005")
    ws.cell(row=7, column=1, value="• Вес: число с точкой или запятой, например 65.5")
    ws.cell(row=8, column=1, value="• Все поля обязательны, кроме 'Тренер'")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="template_participants_{tournament.id}.xlsx"'
    wb.save(response)
    return response


def _parse_date_from_excel(value):
    """Универсальный парсер даты из Excel"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date') and not isinstance(value, str):
        return value.date()
    raw = str(value).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def import_fighters_from_excel(file_path, tournament_id):
    """Импорт бойцов из Excel-файла"""
    tournament = Tournament.objects.get(pk=tournament_id)
    with open(file_path, 'rb') as f:
        file_bytes = io.BytesIO(f.read())
    wb = load_workbook(file_bytes)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or '').strip().lower()
        if 'имя' in val and 'фамилия' not in val:
            headers['first_name'] = col
        elif 'фамилия' in val:
            headers['last_name'] = col
        elif 'дата' in val or 'рождения' in val:
            headers['date_of_birth'] = col
        elif 'пол' in val or 'gender' in val:
            headers['gender'] = col
        elif 'вес' in val:
            headers['weight'] = col
        elif 'клуб' in val:
            headers['club'] = col
        elif 'тренер' in val:
            headers['coach'] = col

    if not headers:
        headers = {
            'first_name': 1, 'last_name': 2, 'date_of_birth': 3,
            'gender': 4, 'weight': 5, 'club': 6, 'coach': 7
        }

    result = {'total': 0, 'success': 0, 'skipped': 0, 'errors': []}
    seen = set()

    for row in range(2, ws.max_row + 1):
        first_name = str(ws.cell(row=row, column=headers.get('first_name', 1)).value or '').strip()
        last_name = str(ws.cell(row=row, column=headers.get('last_name', 2)).value or '').strip()
        dob_val = ws.cell(row=row, column=headers.get('date_of_birth', 3)).value
        gender_raw = str(ws.cell(row=row, column=headers.get('gender', 4)).value or '').strip()
        weight_raw = str(ws.cell(row=row, column=headers.get('weight', 5)).value or '').strip()
        club = str(ws.cell(row=row, column=headers.get('club', 6)).value or '').strip()
        coach = str(ws.cell(row=row, column=headers.get('coach', 7)).value or '').strip()

        if not first_name and not last_name:
            continue

        result['total'] += 1

        file_key = f"{first_name.lower()}|{last_name.lower()}|{str(dob_val)}|{club.lower()}"
        if file_key in seen:
            result['skipped'] += 1
            continue
        seen.add(file_key)

        dob = _parse_date_from_excel(dob_val)
        if dob is None:
            result['errors'].append(f"Строка {row}: неверная дата '{dob_val}' (ожидается ДД.ММ.ГГГГ или ГГГГ-ММ-ДД)")
            continue

        gender = 'M' if gender_raw.lower() in ('м', 'm', 'муж', 'мужской', 'male') else 'F'

        try:
            weight = float(weight_raw.replace(',', '.'))
        except (ValueError, TypeError):
            result['errors'].append(f"Строка {row}: неверный вес '{weight_raw}'")
            continue

        if not first_name or not last_name or not club:
            result['errors'].append(f"Строка {row}: не заполнены обязательные поля")
            continue

        existing = Fighter.objects.filter(
            first_name__iexact=first_name,
            last_name__iexact=last_name,
            date_of_birth=dob,
            club__iexact=club
        ).first()

        if existing:
            fighter = existing
            fighter.weight = weight
            fighter.coach = coach or fighter.coach
            fighter.save()
            result['skipped'] += 1
        else:
            password = f"fighter{dob.strftime('%d%m%y')}"
            temp_username = f"temp_{uuid.uuid4().hex[:12]}"
            user = User.objects.create_user(
                username=temp_username,
                email='',
                first_name=first_name,
                last_name=last_name,
                password=password
            )
            fighter = Fighter.objects.create(
                user=user,
                first_name=first_name,
                last_name=last_name,
                date_of_birth=dob,
                gender=gender,
                weight=weight,
                club=club,
                coach=coach or ''
            )
            user.username = f"fighter{fighter.id}"
            user.save(update_fields=['username'])
            result['success'] += 1

        reg, _ = TournamentRegistration.objects.get_or_create(
            tournament=tournament,
            fighter=fighter,
            defaults={'is_approved': False}
        )

        for cp in tournament.checkpoints.all():
            RegistrationCheckpoint.objects.get_or_create(
                registration=reg, checkpoint=cp, defaults={'is_checked': False}
            )

    wb.close()
    return result


@login_required
@tournament_admin_required
def add_participant_excel(request, tournament_id):
    """Импорт участников из Excel"""
    tournament = get_object_or_404(Tournament, pk=tournament_id)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Неверный формат. Используйте .xlsx или .xls')
                return render(request, 'tournaments/add_participant_excel.html', {
                    'form': form, 'tournament': tournament
                })

            fs = FileSystemStorage(location=tempfile.gettempdir())
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)

            try:
                result = import_fighters_from_excel(file_path, tournament_id)
                if result['success'] > 0:
                    messages.success(request, f"Добавлено {result['success']} из {result['total']} участников")
                if result['skipped'] > 0:
                    messages.info(request, f"Пропущено {result['skipped']} существующих/дублированных")
                for error in result['errors'][:10]:
                    messages.error(request, error)
            except Exception as e:
                messages.error(request, f'Ошибка обработки файла: {e}')
                logger.exception('Excel import error')
            finally:
                try:
                    fs.delete(filename)
                except (PermissionError, OSError):
                    pass

            return redirect(f"{reverse('tournaments:participants_management')}?tournament={tournament_id}")
    else:
        form = ExcelUploadForm()

    return render(request, 'tournaments/add_participant_excel.html', {
        'form': form, 'tournament': tournament
    })


def fighter_list(request):
    """Публичный список бойцов"""
    fighters = Fighter.objects.all().select_related('user')

    search = request.GET.get('search', '')
    gender = request.GET.get('gender', '')
    club = request.GET.get('club', '')

    if search:
        fighters = fighters.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
        )
    if gender:
        fighters = fighters.filter(gender=gender)
    if club:
        fighters = fighters.filter(club__icontains=club)

    return render(request, 'tournaments/fighter_list.html', {
        'fighters': fighters.order_by('last_name', 'first_name'),
        'search': search,
        'gender_filter': gender,
    })


def fighter_detail(request, pk):
    """Профиль бойца"""
    fighter = get_object_or_404(Fighter.objects.select_related('user'), pk=pk)
    return render(request, 'tournaments/fighter_detail.html', {
        'fighter': fighter,
        'stats': fighter.get_stats(),
        'registrations': fighter.registrations.select_related('tournament', 'age_weight_category'),
    })